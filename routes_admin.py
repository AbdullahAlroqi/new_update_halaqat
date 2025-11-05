from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response
from flask_login import login_required, current_user
from models import db, User, Role, LeaveRequest, LeaveType, Schedule, Attendance, SystemSettings, Notification, ActivityLog, AbsenceStatus, Certificate
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl import Workbook
import os
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_RIGHT, TA_CENTER
from sqlalchemy import or_, and_

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

# دالة مساعدة لتسجيل النشاطات
def log_activity(action, target_type, target_id=None, details=None):
    """تسجيل نشاط في سجل النشاطات"""
    try:
        ip_address = request.remote_addr if request else None
        log = ActivityLog(
            user_id=current_user.id,
            action=action,
            target_type=target_type,
            target_id=target_id,
            details=details,
            ip_address=ip_address
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        print(f"Error logging activity: {str(e)}")

# التحقق من صلاحيات الإدارة
def admin_required():
    return current_user.is_authenticated and current_user.role in [Role.MAIN_ADMIN, Role.SUB_ADMIN]

# لوحة تحكم الإدارة
@admin_bp.route('/dashboard')
@login_required
def dashboard():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # إحصائيات عامة
    total_employees = User.query.filter_by(role=Role.EMPLOYEE).count()
    total_supervisors = User.query.filter(
        User.role.in_([Role.MAIN_SUPERVISOR, Role.SUB_SUPERVISOR])
    ).count()
    pending_leaves = LeaveRequest.query.filter_by(status='قيد الانتظار').count()
    
    # الحضور اليوم
    today = datetime.now().date()
    present_today = Attendance.query.filter(
        Attendance.date == today,
        Attendance.status == 'حاضر'
    ).count()
    
    # آخر طلبات الإجازات
    recent_leaves = LeaveRequest.query.order_by(LeaveRequest.created_at.desc()).limit(10).all()
    
    return render_template('admin/dashboard.html',
                         total_employees=total_employees,
                         total_supervisors=total_supervisors,
                         pending_leaves=pending_leaves,
                         present_today=present_today,
                         recent_leaves=recent_leaves)

# إدارة المشرفين
@admin_bp.route('/supervisors')
@login_required
def supervisors():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    supervisors_list = User.query.filter(
        User.role.in_([Role.MAIN_SUPERVISOR, Role.SUB_SUPERVISOR])
    ).all()
    
    return render_template('admin/supervisors.html', supervisors=supervisors_list)

# إضافة مشرف جديد
@admin_bp.route('/supervisors/add', methods=['GET', 'POST'])
@login_required
def add_supervisor():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        national_id = request.form.get('national_id')
        name = request.form.get('name')
        password = request.form.get('password')
        role = request.form.get('role')
        gender = request.form.get('gender')
        department = request.form.get('department')
        shift_start = request.form.get('shift_start')
        shift_end = request.form.get('shift_end')
        shift_time = f"{shift_start} - {shift_end}" if shift_start and shift_end else None
        
        # التحقق من عدم وجود الهوية مسبقاً
        existing = User.query.filter_by(national_id=national_id).first()
        if existing:
            flash('رقم الهوية موجود مسبقاً', 'danger')
            return redirect(url_for('admin.add_supervisor'))
        
        supervisor = User(
            national_id=national_id,
            name=name,
            role=role,
            gender=gender,
            department=department,
            shift_time=shift_time
        )
        supervisor.set_password(password)
        
        db.session.add(supervisor)
        db.session.commit()
        
        # تسجيل النشاط
        log_activity('إضافة', 'مشرف', supervisor.id, f'تم إضافة المشرف: {supervisor.name}')
        
        flash('تم إضافة المشرف بنجاح', 'success')
        return redirect(url_for('admin.supervisors'))
    
    return render_template('admin/add_supervisor.html')

# إدارة الموظفين
@admin_bp.route('/employees')
@login_required
def employees():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # الفلاتر
    gender_filter = request.args.get('gender', '')
    department_filter = request.args.get('department', '')
    name_filter = request.args.get('name', '')
    
    query = User.query.filter_by(role=Role.EMPLOYEE)
    
    if gender_filter:
        query = query.filter_by(gender=gender_filter)
    if department_filter:
        query = query.filter_by(department=department_filter)
    if name_filter:
        query = query.filter(User.name.like(f'%{name_filter}%'))
    
    employees_list = query.all()
    
    # قوائم الفلاتر
    departments = db.session.query(User.department).filter_by(role=Role.EMPLOYEE).distinct().all()
    departments = [d[0] for d in departments if d[0]]
    
    return render_template('admin/employees.html', 
                          employees=employees_list,
                          departments=departments)

# إضافة موظف يدوياً
@admin_bp.route('/employees/add', methods=['GET', 'POST'])
@login_required
def add_employee():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        national_id = request.form.get('national_id')
        name = request.form.get('name')
        gender = request.form.get('gender')
        department = request.form.get('department')
        
        # التحقق من عدم وجود الهوية
        existing = User.query.filter_by(national_id=national_id).first()
        if existing:
            flash('رقم الهوية موجود مسبقاً', 'danger')
            return redirect(url_for('admin.add_employee'))
        
        employee = User(
            national_id=national_id,
            name=name,
            role=Role.EMPLOYEE,
            gender=gender,
            department=department
        )
        # كلمة مرور افتراضية = رقم الهوية
        employee.set_password(national_id)
        
        db.session.add(employee)
        db.session.commit()
        
        # تسجيل النشاط
        log_activity('إضافة', 'موظف', employee.id, f'تم إضافة الموظف: {employee.name}')
        
        flash('تم إضافة الموظف بنجاح', 'success')
        return redirect(url_for('admin.employees'))
    
    return render_template('admin/add_employee.html')

# تعديل موظف
@admin_bp.route('/employees/edit/<int:employee_id>', methods=['GET', 'POST'])
@login_required
def edit_employee(employee_id):
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    employee = User.query.get_or_404(employee_id)
    
    if request.method == 'POST':
        employee.name = request.form.get('name')
        employee.national_id = request.form.get('national_id')
        employee.gender = request.form.get('gender')
        employee.department = request.form.get('department')
        
        shift_start = request.form.get('shift_start')
        shift_end = request.form.get('shift_end')
        if shift_start and shift_end:
            employee.shift_time = f"{shift_start} - {shift_end}"
        
        password = request.form.get('password')
        if password:
            employee.set_password(password)
        
        db.session.commit()
        
        # تسجيل النشاط
        log_activity('تعديل', 'موظف', employee.id, f'تم تعديل معلومات الموظف: {employee.name}')
        
        flash('تم تعديل الموظف بنجاح', 'success')
        return redirect(url_for('admin.employees'))
    
    # تحويل shift_time إلى start و end
    shift_parts = employee.shift_time.split(' - ') if employee.shift_time else ['', '']
    
    return render_template('admin/edit_employee.html', 
                          employee=employee,
                          shift_start=shift_parts[0] if len(shift_parts) > 0 else '',
                          shift_end=shift_parts[1] if len(shift_parts) > 1 else '')

# حذف موظف
@admin_bp.route('/employees/delete/<int:employee_id>', methods=['POST'])
@login_required
def delete_employee(employee_id):
    if not admin_required():
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية لهذه العملية'}), 403
    
    employee = User.query.get_or_404(employee_id)
    
    if employee.role != Role.EMPLOYEE:
        return jsonify({'success': False, 'message': 'هذا المستخدم ليس موظفاً'}), 400
    
    try:
        employee_name = employee.name
        
        # حذف السجلات المرتبطة
        Attendance.query.filter_by(employee_id=employee.id).delete()
        LeaveRequest.query.filter_by(employee_id=employee.id).delete()
        Schedule.query.filter_by(employee_id=employee.id).delete()
        Notification.query.filter_by(user_id=employee.id).delete()
        
        # حذف الموظف
        db.session.delete(employee)
        db.session.commit()
        
        # تسجيل النشاط
        log_activity('حذف', 'موظف', employee_id, f'تم حذف الموظف: {employee_name}')
        
        flash(f'تم حذف الموظف {employee_name} بنجاح', 'success')
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'}), 500

# حذف مشرف
@admin_bp.route('/supervisors/delete/<int:supervisor_id>', methods=['POST'])
@login_required
def delete_supervisor(supervisor_id):
    if not admin_required():
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية لهذه العملية'}), 403
    
    supervisor = User.query.get_or_404(supervisor_id)
    
    if supervisor.role not in [Role.MAIN_SUPERVISOR, Role.SUB_SUPERVISOR]:
        return jsonify({'success': False, 'message': 'هذا المستخدم ليس مشرفاً'}), 400
    
    try:
        supervisor_name = supervisor.name
        
        # إلغاء إسناد الموظفين التابعين له
        subordinates = User.query.filter_by(supervisor_id=supervisor.id).all()
        for emp in subordinates:
            emp.supervisor_id = None
        
        # حذف السجلات المرتبطة
        Notification.query.filter_by(user_id=supervisor.id).delete()
        Schedule.query.filter_by(created_by=supervisor.id).delete()
        
        # حذف المشرف
        db.session.delete(supervisor)
        db.session.commit()
        
        # تسجيل النشاط
        log_activity('حذف', 'مشرف', supervisor_id, f'تم حذف المشرف: {supervisor_name}')
        
        flash(f'تم حذف المشرف {supervisor_name} بنجاح', 'success')
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'حدث خطأ: {str(e)}'}), 500

# تعديل مشرف
@admin_bp.route('/supervisors/edit/<int:supervisor_id>', methods=['GET', 'POST'])
@login_required
def edit_supervisor(supervisor_id):
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    supervisor = User.query.get_or_404(supervisor_id)
    
    if request.method == 'POST':
        supervisor.name = request.form.get('name')
        supervisor.national_id = request.form.get('national_id')
        supervisor.gender = request.form.get('gender')
        supervisor.department = request.form.get('department')
        supervisor.role = request.form.get('role')
        
        password = request.form.get('password')
        if password:
            supervisor.set_password(password)
        
        db.session.commit()
        
        # تسجيل النشاط
        log_activity('تعديل', 'مشرف', supervisor.id, f'تم تعديل معلومات المشرف: {supervisor.name}')
        
        flash('تم تعديل المشرف بنجاح', 'success')
        return redirect(url_for('admin.supervisors'))
    
    return render_template('admin/edit_supervisor.html', supervisor=supervisor)

# رفع ملف Excel للموظفين
@admin_bp.route('/employees/upload', methods=['GET', 'POST'])
@login_required
def upload_employees():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(url_for('admin.upload_employees'))
        
        file = request.files['file']
        if file.filename == '':
            flash('لم يتم اختيار ملف', 'danger')
            return redirect(url_for('admin.upload_employees'))
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                wb = openpyxl.load_workbook(file)
                sheet = wb.active
                
                added_count = 0
                skipped_count = 0
                
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # التحقق من وجود بيانات
                    if not row or not row[0]:
                        continue
                    
                    # قراءة البيانات بالترتيب الصحيح (حسب الصورة)
                    name = str(row[0]).strip() if row[0] else ''  # الاسم
                    national_id = str(row[1]).strip() if row[1] else ''  # الهوية
                    period = str(row[2]).strip() if row[2] else ''  # الفترة
                    work_time = str(row[3]).strip() if row[3] else ''  # الوقت
                    rest_days = str(row[4]).strip() if row[4] else ''  # أيام الراحة
                    department = str(row[5]).strip() if len(row) > 5 and row[5] else 'الحلقات'  # القسم
                    gender = str(row[6]).strip() if len(row) > 6 and row[6] else 'ذكر'  # الجنس
                    
                    # التحقق من وجود الموظف
                    existing = User.query.filter_by(national_id=national_id).first()
                    if existing:
                        # تحديث معلومات الجدول للموظف الموجود
                        existing.period = period
                        existing.work_time = work_time
                        existing.rest_days = rest_days
                        existing.department = department
                        existing.gender = gender
                        skipped_count += 1
                        continue
                    
                    employee = User(
                        national_id=national_id,
                        name=name,
                        role=Role.EMPLOYEE,
                        gender=gender,
                        department=department,
                        period=period,
                        work_time=work_time,
                        rest_days=rest_days
                    )
                    employee.set_password(national_id)
                    
                    db.session.add(employee)
                    added_count += 1
                
                db.session.commit()
                flash(f'تم إضافة {added_count} موظف. تم تخطي {skipped_count} موظف موجود مسبقاً', 'success')
                
            except Exception as e:
                flash(f'حدث خطأ أثناء قراءة الملف: {str(e)}', 'danger')
        else:
            flash('نوع الملف غير مدعوم. الرجاء رفع ملف Excel', 'danger')
        
        return redirect(url_for('admin.employees'))
    
    return render_template('admin/upload_employees.html')

# تحميل نموذج Excel
@admin_bp.route('/employees/download-template')
@login_required
def download_template():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    wb = Workbook()
    sheet = wb.active
    sheet.title = "الموظفين"
    
    # العناوين بالترتيب الصحيح
    headers = ['الاسم', 'الهوية', 'الفترة', 'الوقت', 'الراحة', 'القسم', 'الجنس']
    sheet.append(headers)
    
    # حفظ في الذاكرة
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='نموذج_الموظفين.xlsx'
    )

# إسناد الموظفين للمشرفين
@admin_bp.route('/assign-employees', methods=['GET', 'POST'])
@login_required
def assign_employees():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    supervisors = User.query.filter(
        User.role.in_([Role.MAIN_SUPERVISOR, Role.SUB_SUPERVISOR])
    ).all()
    
    employees = User.query.filter_by(role=Role.EMPLOYEE).all()
    
    if request.method == 'POST':
        supervisor_id = request.form.get('supervisor_id', type=int)
        employee_ids = request.form.getlist('employee_ids')
        
        for emp_id in employee_ids:
            employee = User.query.get(int(emp_id))
            if employee:
                employee.supervisor_id = supervisor_id
        
        db.session.commit()
        flash('تم إسناد الموظفين بنجاح', 'success')
        return redirect(url_for('admin.assign_employees'))
    
    return render_template('admin/assign_employees.html', 
                         supervisors=supervisors,
                         employees=employees)

# إدارة أنواع الإجازات
@admin_bp.route('/leave-types')
@login_required
def leave_types():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    leave_types_list = LeaveType.query.all()
    return render_template('admin/leave_types.html', leave_types=leave_types_list)

# إضافة نوع إجازة
@admin_bp.route('/leave-types/add', methods=['POST'])
@login_required
def add_leave_type():
    if not admin_required():
        return jsonify({'success': False}), 403
    
    name = request.form.get('name')
    max_days = request.form.get('max_days', type=int)
    requires_attachment = request.form.get('requires_attachment') == 'on'
    deduct_from_balance = request.form.get('deduct_from_balance') == 'on'
    
    leave_type = LeaveType(
        name=name,
        max_days=max_days,
        requires_attachment=requires_attachment,
        deduct_from_balance=deduct_from_balance
    )
    
    db.session.add(leave_type)
    db.session.commit()
    
    flash('تم إضافة نوع الإجازة بنجاح', 'success')
    return redirect(url_for('admin.leave_types'))

# تعديل نوع إجازة
@admin_bp.route('/leave-types/edit/<int:leave_type_id>', methods=['POST'])
@login_required
def edit_leave_type(leave_type_id):
    if not admin_required():
        return jsonify({'success': False}), 403
    
    leave_type = LeaveType.query.get_or_404(leave_type_id)
    
    leave_type.name = request.form.get('name')
    leave_type.max_days = request.form.get('max_days', type=int)
    leave_type.requires_attachment = request.form.get('requires_attachment') == 'on'
    leave_type.deduct_from_balance = request.form.get('deduct_from_balance') == 'on'
    leave_type.is_active = request.form.get('is_active') == 'on'
    
    db.session.commit()
    
    flash('تم تعديل نوع الإجازة بنجاح', 'success')
    return redirect(url_for('admin.leave_types'))

# حذف نوع إجازة
@admin_bp.route('/leave-types/delete/<int:leave_type_id>', methods=['POST'])
@login_required
def delete_leave_type(leave_type_id):
    if not admin_required():
        return jsonify({'success': False}), 403
    
    leave_type = LeaveType.query.get_or_404(leave_type_id)
    
    # التحقق من عدم وجود طلبات إجازة مرتبطة بهذا النوع
    if leave_type.leave_requests.count() > 0:
        flash('لا يمكن حذف نوع الإجازة لأنه مرتبط بطلبات إجازة', 'danger')
        return redirect(url_for('admin.leave_types'))
    
    db.session.delete(leave_type)
    db.session.commit()
    
    flash('تم حذف نوع الإجازة بنجاح', 'success')
    return redirect(url_for('admin.leave_types'))

# إعدادات النظام
@admin_bp.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if current_user.role != Role.MAIN_ADMIN:
        flash('هذه الصفحة متاحة فقط لمدير النظام الأساسي', 'danger')
        return redirect(url_for('index'))
    
    system_settings = SystemSettings.query.first()
    if not system_settings:
        system_settings = SystemSettings()
        db.session.add(system_settings)
        db.session.commit()
    
    if request.method == 'POST':
        system_settings.system_name = request.form.get('system_name')
        system_settings.primary_color = request.form.get('primary_color')
        system_settings.secondary_color = request.form.get('secondary_color')
        system_settings.accent_color = request.form.get('accent_color')
        system_settings.attachment_retention_days = request.form.get('attachment_retention_days', type=int)
        
        db.session.commit()
        flash('تم تحديث الإعدادات بنجاح', 'success')
        return redirect(url_for('admin.settings'))
    
    return render_template('admin/settings.html', settings=system_settings)

# التقارير
@admin_bp.route('/reports')
@login_required
def reports():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    return render_template('admin/reports.html')

# تقرير الإجازات
@admin_bp.route('/reports/leaves')
@login_required
def report_leaves():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status = request.args.get('status')
    
    query = LeaveRequest.query
    
    if start_date:
        query = query.filter(LeaveRequest.start_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(LeaveRequest.end_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if status:
        query = query.filter(LeaveRequest.status == status)
    
    leaves = query.order_by(LeaveRequest.created_at.desc()).all()
    
    return render_template('admin/report_leaves.html', leaves=leaves)

# تقرير الحضور
@admin_bp.route('/reports/attendance')
@login_required
def report_attendance():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    employee_id = request.args.get('employee_id', type=int)
    
    query = Attendance.query
    
    if start_date:
        query = query.filter(Attendance.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Attendance.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if employee_id:
        query = query.filter(Attendance.employee_id == employee_id)
    
    records = query.order_by(Attendance.date.desc()).all()
    employees = User.query.filter_by(role=Role.EMPLOYEE).all()
    
    return render_template('admin/report_attendance.html', records=records, employees=employees)

# تخصيص المظهر (الألوان والشعار)
@admin_bp.route('/customize', methods=['GET', 'POST'])
@login_required
def customize():
    if current_user.role != Role.MAIN_ADMIN:
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    settings = SystemSettings.query.first()
    if not settings:
        settings = SystemSettings()
        db.session.add(settings)
        db.session.commit()
    
    if request.method == 'POST':
        # تحديث الألوان
        settings.primary_color = request.form.get('primary_color', '#0d7377')
        settings.secondary_color = request.form.get('secondary_color', '#14FFEC')
        settings.accent_color = request.form.get('accent_color', '#323232')
        settings.system_name = request.form.get('system_name', 'نظام إدارة معلمي الحلقات - مكة المكرمة')
        
        # رفع الشعار
        if 'logo' in request.files:
            logo = request.files['logo']
            if logo and logo.filename:
                filename = secure_filename('logo_' + logo.filename)
                logo_folder = os.path.join('static', 'images')
                os.makedirs(logo_folder, exist_ok=True)
                logo_path = os.path.join(logo_folder, filename)
                logo.save(logo_path)
                settings.logo_path = filename  # نحفظ اسم الملف فقط
        
        db.session.commit()
        flash('تم تحديث التخصيصات بنجاح', 'success')
        return redirect(url_for('admin.customize'))
    
    return render_template('admin/customize.html', settings=settings)

# طباعة تقرير الإجازات PDF
@admin_bp.route('/reports/leaves/pdf')
@login_required
def report_leaves_pdf():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # جلب البيانات
    employee_id = request.args.get('employee_id', type=int)
    leave_type_id = request.args.get('leave_type_id', type=int)
    status = request.args.get('status')
    
    # تحديد الـ join بشكل صريح لتجنب AmbiguousForeignKeysError
    query = LeaveRequest.query.join(User, LeaveRequest.employee_id == User.id).join(LeaveType)
    
    if employee_id:
        query = query.filter(LeaveRequest.employee_id == employee_id)
    if leave_type_id:
        query = query.filter(LeaveRequest.leave_type_id == leave_type_id)
    if status:
        query = query.filter(LeaveRequest.status == status)
    
    leaves = query.order_by(LeaveRequest.created_at.desc()).all()
    
    # إنشاء PDF مع دعم العربية
    from reportlab.pdfbase.pdfmetrics import registerFont
    from reportlab.pdfbase.ttfonts import TTFont
    from arabic_reshaper import reshape
    from bidi.algorithm import get_display
    
    # تسجيل خط عربي (استخدام Arial Unicode MS)
    try:
        registerFont(TTFont('Arabic', 'C:/Windows/Fonts/arial.ttf'))
        arabic_font = 'Arabic'
    except:
        # إذا فشل، نستخدم الخط الافتراضي
        arabic_font = 'Helvetica'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    
    # العنوان بالعربية
    title_text = 'تقرير الإجازات'
    reshaped_title = reshape(title_text)
    bidi_title = get_display(reshaped_title)
    
    title_style = ParagraphStyle(
        'TitleStyle',
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20,
        fontName=arabic_font
    )
    elements.append(Paragraph(bidi_title, title_style))
    elements.append(Spacer(1, 20))
    
    # الجدول
    def arabic_text(text):
        """تحويل النص العربي للعرض الصحيح"""
        reshaped = reshape(text)
        return get_display(reshaped)
    
    headers = ['الموظف', 'نوع الإجازة', 'من تاريخ', 'إلى تاريخ', 'الأيام', 'الحالة']
    data = [[arabic_text(h) for h in headers]]
    
    for leave in leaves:
        data.append([
            arabic_text(leave.employee.name),
            arabic_text(leave.leave_type.name),
            leave.start_date.strftime('%Y-%m-%d'),
            leave.end_date.strftime('%Y-%m-%d'),
            str(leave.days_count),
            arabic_text(leave.status)
        ])
    
    table = Table(data, colWidths=[4*cm, 3*cm, 2.5*cm, 2.5*cm, 2*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d7377')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), arabic_font),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'leaves_report_{datetime.now().strftime("%Y%m%d")}.pdf', mimetype='application/pdf')

# طباعة تقرير الحضور PDF
@admin_bp.route('/reports/attendance/pdf')
@login_required  
def report_attendance_pdf():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # جلب البيانات
    employee_id = request.args.get('employee_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = Attendance.query.join(User, Attendance.employee_id == User.id)
    
    if start_date:
        query = query.filter(Attendance.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Attendance.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if employee_id:
        query = query.filter(Attendance.employee_id == employee_id)
    
    records = query.order_by(Attendance.date.desc()).all()
    
    # إنشاء PDF مع دعم العربية
    from reportlab.pdfbase.pdfmetrics import registerFont
    from reportlab.pdfbase.ttfonts import TTFont
    from arabic_reshaper import reshape
    from bidi.algorithm import get_display
    
    # تسجيل خط عربي
    try:
        registerFont(TTFont('Arabic', 'C:/Windows/Fonts/arial.ttf'))
        arabic_font = 'Arabic'
    except:
        arabic_font = 'Helvetica'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    
    # العنوان بالعربية
    title_text = 'تقرير الحضور والغياب'
    reshaped_title = reshape(title_text)
    bidi_title = get_display(reshaped_title)
    
    title_style = ParagraphStyle(
        'TitleStyle',
        fontSize=18,
        alignment=TA_CENTER,
        spaceAfter=20,
        fontName=arabic_font
    )
    elements.append(Paragraph(bidi_title, title_style))
    elements.append(Spacer(1, 20))
    
    # الجدول
    def arabic_text(text):
        """تحويل النص العربي للعرض الصحيح"""
        reshaped = reshape(text)
        return get_display(reshaped)
    
    headers = ['الموظف', 'التاريخ', 'الحالة', 'الملاحظات']
    data = [[arabic_text(h) for h in headers]]
    
    for record in records:
        data.append([
            arabic_text(record.employee.name),
            record.date.strftime('%Y-%m-%d'),
            arabic_text(record.status),
            arabic_text(record.notes) if record.notes else '-'
        ])
    
    table = Table(data, colWidths=[5*cm, 3*cm, 3*cm, 5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d7377')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), arabic_font),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f'attendance_report_{datetime.now().strftime("%Y%m%d")}.pdf', mimetype='application/pdf')

# حذف بيانات الاختبار (مدير أساسي فقط)
@admin_bp.route('/delete-test-data', methods=['POST'])
@login_required
def delete_test_data_route():
    if current_user.role != Role.MAIN_ADMIN:
        flash('ليس لديك صلاحية لهذه العملية', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    try:
        # حذف سجلات الحضور للموظفين التجريبيين
        test_employees = User.query.filter(
            User.national_id.like('4000%') | User.national_id.like('5000%')
        ).all()
        
        for emp in test_employees:
            Attendance.query.filter_by(employee_id=emp.id).delete()
            LeaveRequest.query.filter_by(employee_id=emp.id).delete()
            Schedule.query.filter_by(employee_id=emp.id).delete()
        
        # حذف المستخدمين التجريبيين
        User.query.filter(
            User.national_id.like('4000%') | User.national_id.like('5000%')
        ).delete(synchronize_session=False)
        
        User.query.filter(
            User.national_id.like('2000%') | User.national_id.like('3000%')
        ).delete(synchronize_session=False)
        
        db.session.commit()
        flash('تم حذف بيانات الاختبار بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('admin.dashboard'))

# عرض طلبات الإجازات للإدارة
@admin_bp.route('/leave-requests')
@login_required
def leave_requests():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # جلب جميع طلبات الإجازات
    status_filter = request.args.get('status', 'قيد الانتظار')
    
    query = LeaveRequest.query.join(User, LeaveRequest.employee_id == User.id).join(LeaveType)
    
    if status_filter and status_filter != 'all':
        query = query.filter(LeaveRequest.status == status_filter)
    
    requests = query.order_by(LeaveRequest.created_at.desc()).all()
    
    return render_template('admin/leave_requests.html', requests=requests, status_filter=status_filter)

# مراجعة طلب إجازة (قبول/رفض)
@admin_bp.route('/review-leave/<int:request_id>', methods=['POST'])
@login_required
def review_leave(request_id):
    if not admin_required():
        flash('ليس لديك صلاحية لهذه العملية', 'danger')
        return redirect(url_for('index'))
    
    leave_request = LeaveRequest.query.get_or_404(request_id)
    action = request.form.get('action')
    notes = request.form.get('notes', '')
    
    if action == 'approve':
        leave_request.status = 'مقبول'
        leave_request.reviewed_by = current_user.id
        leave_request.reviewed_at = datetime.now()
        leave_request.review_notes = notes
        
        # خصم من رصيد الإجازات إذا كان نوع الإجازة يتطلب ذلك
        if leave_request.leave_type.deduct_from_balance:
            employee = leave_request.employee
            employee.leave_balance -= leave_request.days_count
            
            # تسجيل النشاط
            log_activity('خصم رصيد إجازة', 'user', employee.id, 
                        f'تم خصم {leave_request.days_count} يوم من رصيد {employee.name} (نوع: {leave_request.leave_type.name})')
        
        flash('تم قبول طلب الإجازة بنجاح', 'success')
    elif action == 'reject':
        leave_request.status = 'مرفوض'
        leave_request.reviewed_by = current_user.id
        leave_request.reviewed_at = datetime.now()
        leave_request.review_notes = notes
        flash('تم رفض طلب الإجازة', 'warning')
    
    db.session.commit()
    
    return redirect(url_for('admin.leave_requests'))

# عرض جدول الحلقات
@admin_bp.route('/schedules-table')
@login_required
def schedules_table():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # الفلاتر
    gender_filter = request.args.get('gender', '')
    department_filter = request.args.get('department', '')
    period_filter = request.args.get('period', '')
    name_filter = request.args.get('name', '')
    
    # جلب جميع المعلمين (الموظفين) مع معلومات جدولهم
    query = User.query.filter_by(role=Role.EMPLOYEE, is_active=True)
    
    # تطبيق الفلاتر
    if gender_filter:
        query = query.filter_by(gender=gender_filter)
    if department_filter:
        query = query.filter_by(department=department_filter)
    if period_filter:
        query = query.filter_by(period=period_filter)
    if name_filter:
        query = query.filter(User.name.like(f'%{name_filter}%'))
    
    # ترتيب حسب الفترة والوقت ثم الاسم
    employees = query.order_by(User.period, User.work_time, User.name).all()
    
    # قوائم الفلاتر
    departments = db.session.query(User.department).filter_by(role=Role.EMPLOYEE).distinct().all()
    departments = [d[0] for d in departments if d[0]]
    
    periods = db.session.query(User.period).filter_by(role=Role.EMPLOYEE).distinct().all()
    periods = [p[0] for p in periods if p[0]]
    
    return render_template('admin/schedules_table.html', 
                          employees=employees,
                          departments=departments,
                          periods=periods)

# حذف جميع بيانات المعلمين
@admin_bp.route('/delete-all-employees', methods=['POST'])
@login_required
def delete_all_employees():
    if current_user.role != Role.MAIN_ADMIN:
        flash('ليس لديك صلاحية لهذه العملية', 'danger')
        return redirect(url_for('admin.schedules_table'))
    
    try:
        # حذف جميع سجلات المعلمين
        employees = User.query.filter_by(role=Role.EMPLOYEE).all()
        
        for emp in employees:
            # حذف السجلات المرتبطة
            Attendance.query.filter_by(employee_id=emp.id).delete()
            LeaveRequest.query.filter_by(employee_id=emp.id).delete()
            Schedule.query.filter_by(employee_id=emp.id).delete()
            
            # حذف الإشعارات المرتبطة
            from models import Notification
            Notification.query.filter_by(user_id=emp.id).delete()
            
            # حذف الموظف
            db.session.delete(emp)
        
        db.session.commit()
        flash(f'تم حذف جميع بيانات المعلمين بنجاح ({len(employees)} معلم)', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ: {str(e)}', 'danger')
    
    return redirect(url_for('admin.schedules_table'))

# صفحة إعدادات الحساب
@admin_bp.route('/account-settings', methods=['GET', 'POST'])
@login_required
def account_settings():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        # تغيير كلمة السر
        if action == 'change_password':
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            # التحقق من كلمة السر الحالية
            if not current_user.check_password(current_password):
                flash('كلمة السر الحالية غير صحيحة', 'danger')
                return redirect(url_for('admin.account_settings'))
            
            # التحقق من تطابق كلمة السر الجديدة
            if new_password != confirm_password:
                flash('كلمة السر الجديدة غير متطابقة', 'danger')
                return redirect(url_for('admin.account_settings'))
            
            # التحقق من طول كلمة السر
            if len(new_password) < 6:
                flash('كلمة السر يجب أن تكون 6 أحرف على الأقل', 'danger')
                return redirect(url_for('admin.account_settings'))
            
            # تغيير كلمة السر
            current_user.set_password(new_password)
            db.session.commit()
            flash('تم تغيير كلمة السر بنجاح', 'success')
            return redirect(url_for('admin.account_settings'))
        
        # تغيير رقم الهوية
        elif action == 'change_national_id':
            new_national_id = request.form.get('new_national_id')
            password_confirm = request.form.get('password_confirm')
            
            # التحقق من كلمة السر
            if not current_user.check_password(password_confirm):
                flash('كلمة السر غير صحيحة', 'danger')
                return redirect(url_for('admin.account_settings'))
            
            # التحقق من صحة رقم الهوية
            if len(new_national_id) != 10 or not new_national_id.isdigit():
                flash('رقم الهوية يجب أن يكون 10 أرقام', 'danger')
                return redirect(url_for('admin.account_settings'))
            
            # التحقق من عدم تكرار رقم الهوية
            existing = User.query.filter_by(national_id=new_national_id).first()
            if existing and existing.id != current_user.id:
                flash('رقم الهوية موجود لمستخدم آخر', 'danger')
                return redirect(url_for('admin.account_settings'))
            
            # تغيير رقم الهوية
            old_id = current_user.national_id
            current_user.national_id = new_national_id
            db.session.commit()
            flash(f'تم تغيير رقم الهوية من {old_id} إلى {new_national_id}', 'success')
            return redirect(url_for('admin.account_settings'))
    
    return render_template('admin/account_settings.html')

# إدارة مديري النظام
@admin_bp.route('/system-admins')
@login_required
def system_admins():
    if current_user.role != Role.MAIN_ADMIN:
        flash('هذه الصفحة متاحة فقط لمدير النظام الأساسي', 'danger')
        return redirect(url_for('admin.dashboard'))
    
    admins = User.query.filter(User.role.in_([Role.MAIN_ADMIN, Role.SUB_ADMIN])).all()
    return render_template('admin/system_admins.html', admins=admins)

# إضافة مدير نظام فرعي
@admin_bp.route('/system-admins/add', methods=['POST'])
@login_required
def add_system_admin():
    if current_user.role != Role.MAIN_ADMIN:
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية لهذه العملية'}), 403
    
    national_id = request.form.get('national_id')
    name = request.form.get('name')
    password = request.form.get('password')
    gender = request.form.get('gender')
    
    # التحقق من عدم وجود الهوية
    existing = User.query.filter_by(national_id=national_id).first()
    if existing:
        return jsonify({'success': False, 'message': 'رقم الهوية موجود مسبقاً'}), 400
    
    admin = User(
        national_id=national_id,
        name=name,
        role=Role.SUB_ADMIN,
        gender=gender
    )
    admin.set_password(password)
    
    db.session.add(admin)
    db.session.commit()
    
    # تسجيل النشاط
    log_activity('إضافة', 'مدير نظام', admin.id, f'تم إضافة مدير نظام: {admin.name}')
    
    return jsonify({'success': True, 'message': 'تم إضافة مدير النظام بنجاح'})

# تعديل مدير نظام
@admin_bp.route('/system-admins/edit/<int:admin_id>', methods=['POST'])
@login_required
def edit_system_admin(admin_id):
    if current_user.role != Role.MAIN_ADMIN:
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية لهذه العملية'}), 403
    
    admin = User.query.get_or_404(admin_id)
    
    admin.name = request.form.get('name')
    admin.national_id = request.form.get('national_id')
    admin.gender = request.form.get('gender')
    
    password = request.form.get('password')
    if password:
        admin.set_password(password)
    
    db.session.commit()
    
    # تسجيل النشاط
    log_activity('تعديل', 'مدير نظام', admin.id, f'تم تعديل معلومات مدير النظام: {admin.name}')
    
    return jsonify({'success': True, 'message': 'تم تعديل مدير النظام بنجاح'})

# حذف مدير نظام
@admin_bp.route('/system-admins/delete/<int:admin_id>', methods=['POST'])
@login_required
def delete_system_admin(admin_id):
    if current_user.role != Role.MAIN_ADMIN:
        return jsonify({'success': False, 'message': 'ليس لديك صلاحية لهذه العملية'}), 403
    
    # التحقق من عدم حذف الحساب الحالي
    if admin_id == current_user.id:
        return jsonify({'success': False, 'message': 'لا يمكنك حذف حسابك الخاص'}), 400
    
    admin = User.query.get_or_404(admin_id)
    
    if admin.role == Role.MAIN_ADMIN:
        return jsonify({'success': False, 'message': 'لا يمكن حذف مدير النظام الأساسي'}), 400
    
    admin_name = admin.name
    db.session.delete(admin)
    db.session.commit()
    
    # تسجيل النشاط
    log_activity('حذف', 'مدير نظام', admin_id, f'تم حذف مدير النظام: {admin_name}')
    
    return jsonify({'success': True, 'message': f'تم حذف مدير النظام {admin_name} بنجاح'})

# سجل النشاطات (Activity Logs)
@admin_bp.route('/activity-logs')
@login_required
def activity_logs():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # الفلاتر
    action_filter = request.args.get('action')
    target_type_filter = request.args.get('target_type')
    user_id_filter = request.args.get('user_id', type=int)
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = ActivityLog.query.join(User, ActivityLog.user_id == User.id)
    
    if action_filter:
        query = query.filter(ActivityLog.action == action_filter)
    if target_type_filter:
        query = query.filter(ActivityLog.target_type == target_type_filter)
    if user_id_filter:
        query = query.filter(ActivityLog.user_id == user_id_filter)
    if date_from:
        query = query.filter(ActivityLog.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        query = query.filter(ActivityLog.created_at <= datetime.strptime(date_to, '%Y-%m-%d'))
    
    logs = query.order_by(ActivityLog.created_at.desc()).paginate(page=request.args.get('page', 1, type=int), per_page=50)
    
    # قائمة المستخدمين للفلتر
    users = User.query.filter(User.role.in_([Role.MAIN_ADMIN, Role.SUB_ADMIN])).all()
    
    return render_template('admin/activity_logs.html', logs=logs, users=users)

# إدارة حالات الغياب
@admin_bp.route('/absence-statuses')
@login_required
def absence_statuses():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    statuses = AbsenceStatus.query.all()
    return render_template('admin/absence_statuses.html', statuses=statuses)

# إضافة حالة غياب
@admin_bp.route('/absence-statuses/add', methods=['POST'])
@login_required
def add_absence_status():
    if not admin_required():
        return jsonify({'success': False}), 403
    
    name = request.form.get('name')
    color = request.form.get('color')
    is_counted_as_absent = request.form.get('is_counted_as_absent') == 'on'
    
    # التحقق من عدم التكرار
    existing = AbsenceStatus.query.filter_by(name=name).first()
    if existing:
        return jsonify({'success': False, 'message': 'هذه الحالة موجودة مسبقاً'}), 400
    
    status = AbsenceStatus(
        name=name,
        color=color,
        is_counted_as_absent=is_counted_as_absent
    )
    
    db.session.add(status)
    db.session.commit()
    
    # تسجيل النشاط
    log_activity('إضافة', 'حالة غياب', status.id, f'تم إضافة حالة غياب: {status.name}')
    
    return jsonify({'success': True})

# تعديل حالة غياب
@admin_bp.route('/absence-statuses/edit/<int:status_id>', methods=['POST'])
@login_required
def edit_absence_status(status_id):
    if not admin_required():
        return jsonify({'success': False}), 403
    
    status = AbsenceStatus.query.get_or_404(status_id)
    
    status.name = request.form.get('name')
    status.color = request.form.get('color')
    status.is_counted_as_absent = request.form.get('is_counted_as_absent') == 'on'
    status.is_active = request.form.get('is_active') == 'on'
    
    db.session.commit()
    
    # تسجيل النشاط
    log_activity('تعديل', 'حالة غياب', status.id, f'تم تعديل حالة الغياب: {status.name}')
    
    return jsonify({'success': True})

# حذف حالة غياب
@admin_bp.route('/absence-statuses/delete/<int:status_id>', methods=['POST'])
@login_required
def delete_absence_status(status_id):
    if not admin_required():
        return jsonify({'success': False}), 403
    
    status = AbsenceStatus.query.get_or_404(status_id)
    
    # التحقق من عدم وجود سجلات مرتبطة
    if status.attendance_records.count() > 0:
        return jsonify({'success': False, 'message': 'لا يمكن حذف هذه الحالة لأنها مرتبطة بسجلات حضور'}), 400
    
    status_name = status.name
    db.session.delete(status)
    db.session.commit()
    
    # تسجيل النشاط
    log_activity('حذف', 'حالة غياب', status_id, f'تم حذف حالة الغياب: {status_name}')
    
    return jsonify({'success': True})

# صفحة التحضير (الحضور والغياب)
@admin_bp.route('/attendance-management')
@login_required
def attendance_management():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # الفلاتر
    date_filter = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    name_filter = request.args.get('name', '')
    gender_filter = request.args.get('gender', '')
    department_filter = request.args.get('department', '')
    period_filter = request.args.get('period', '')
    
    # بناء الاستعلام
    query = User.query.filter_by(role=Role.EMPLOYEE, is_active=True)
    
    if name_filter:
        query = query.filter(User.name.like(f'%{name_filter}%'))
    if gender_filter:
        query = query.filter_by(gender=gender_filter)
    if department_filter:
        query = query.filter_by(department=department_filter)
    if period_filter:
        query = query.filter_by(period=period_filter)
    
    employees = query.order_by(User.name).all()
    
    # جلب سجلات الحضور لهذا اليوم
    date_obj = datetime.strptime(date_filter, '%Y-%m-%d').date()
    attendance_records = {}
    for emp in employees:
        record = Attendance.query.filter_by(employee_id=emp.id, date=date_obj).first()
        attendance_records[emp.id] = record
    
    # جلب حالات الغياب
    absence_statuses = AbsenceStatus.query.filter_by(is_active=True).all()
    
    # قوائم الفلاتر
    departments = db.session.query(User.department).filter_by(role=Role.EMPLOYEE).distinct().all()
    departments = [d[0] for d in departments if d[0]]
    
    periods = db.session.query(User.period).filter_by(role=Role.EMPLOYEE).distinct().all()
    periods = [p[0] for p in periods if p[0]]
    
    return render_template('admin/attendance_management.html',
                         employees=employees,
                         attendance_records=attendance_records,
                         absence_statuses=absence_statuses,
                         date_filter=date_filter,
                         departments=departments,
                         periods=periods)

# تسجيل الحضور/الغياب
@admin_bp.route('/mark-attendance', methods=['POST'])
@login_required
def mark_attendance():
    if not admin_required():
        return jsonify({'success': False}), 403
    
    data = request.get_json()
    employee_id = data.get('employee_id')
    date_str = data.get('date')
    status = data.get('status')
    absence_status_id = data.get('absence_status_id')
    notes = data.get('notes', '')
    
    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    
    # البحث عن سجل موجود
    record = Attendance.query.filter_by(employee_id=employee_id, date=date_obj).first()
    
    if record:
        # تحديث السجل
        record.status = status
        record.absence_status_id = absence_status_id
        record.notes = notes
        record.recorded_by = current_user.id
    else:
        # إنشاء سجل جديد
        record = Attendance(
            employee_id=employee_id,
            date=date_obj,
            status=status,
            absence_status_id=absence_status_id,
            notes=notes,
            recorded_by=current_user.id
        )
        db.session.add(record)
    
    db.session.commit()
    
    # تسجيل النشاط
    employee = User.query.get(employee_id)
    log_activity('تحديث حضور', 'موظف', employee_id, f'تم تحديث حضور {employee.name} بتاريخ {date_str}: {status}')
    
    return jsonify({'success': True})

# تعديل جدول موظف
@admin_bp.route('/employees/<int:employee_id>/edit-schedule', methods=['GET', 'POST'])
@login_required
def edit_employee_schedule(employee_id):
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    employee = User.query.get_or_404(employee_id)
    
    if request.method == 'POST':
        employee.period = request.form.get('period')
        employee.work_time = request.form.get('work_time')
        employee.rest_days = request.form.get('rest_days')
        
        db.session.commit()
        
        # تسجيل النشاط
        log_activity('تعديل جدول', 'موظف', employee.id, f'تم تعديل جدول الموظف: {employee.name}')
        
        flash('تم تعديل جدول الموظف بنجاح', 'success')
        return redirect(url_for('admin.schedules_table'))
    
    return render_template('admin/edit_employee_schedule.html', employee=employee)

# تقرير PDF للجدول
@admin_bp.route('/schedules-table/pdf')
@login_required
def schedules_table_pdf():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    # الفلاتر
    gender_filter = request.args.get('gender', '')
    department_filter = request.args.get('department', '')
    period_filter = request.args.get('period', '')
    
    query = User.query.filter_by(role=Role.EMPLOYEE, is_active=True)
    
    if gender_filter:
        query = query.filter_by(gender=gender_filter)
    if department_filter:
        query = query.filter_by(department=department_filter)
    if period_filter:
        query = query.filter_by(period=period_filter)
    
    # ترتيب حسب الفترة والوقت
    employees = query.order_by(User.period, User.work_time, User.name).all()
    
    # إنشاء PDF مع دعم العربية
    from reportlab.pdfbase.pdfmetrics import registerFont
    from reportlab.pdfbase.ttfonts import TTFont
    from arabic_reshaper import reshape
    from bidi.algorithm import get_display
    
    # تسجيل خط عربي
    try:
        registerFont(TTFont('Arabic', 'C:/Windows/Fonts/arial.ttf'))
        arabic_font = 'Arabic'
    except:
        try:
            registerFont(TTFont('Arabic', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
            arabic_font = 'Arabic'
        except:
            arabic_font = 'Helvetica'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    elements = []
    
    # العنوان
    title_text = 'جدول معلمي الحلقات'
    reshaped_title = reshape(title_text)
    bidi_title = get_display(reshaped_title)
    
    title_style = ParagraphStyle(
        'TitleStyle',
        fontSize=20,
        alignment=TA_CENTER,
        spaceAfter=30,
        fontName=arabic_font
    )
    elements.append(Paragraph(bidi_title, title_style))
    
    # دالة لتحويل النص العربي
    def arabic_text(text):
        if not text:
            return '-'
        reshaped = reshape(str(text))
        return get_display(reshaped)
    
    # بيانات الجدول
    headers = ['م', 'الاسم', 'الفترة', 'الوقت', 'أيام الراحة', 'القسم', 'الجنس']
    data = [[arabic_text(h) for h in headers]]
    
    for idx, emp in enumerate(employees, 1):
        data.append([
            str(idx),
            arabic_text(emp.name),
            arabic_text(emp.period),
            arabic_text(emp.work_time),
            arabic_text(emp.rest_days),
            arabic_text(emp.department),
            arabic_text(emp.gender)
        ])
    
    # إنشاء الجدول
    table = Table(data, colWidths=[1.5*cm, 5*cm, 3*cm, 3*cm, 4*cm, 3*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d7377')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), arabic_font),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # التذييل
    footer_text = f'تاريخ الطباعة: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    footer_style = ParagraphStyle(
        'FooterStyle',
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=10,
        fontName=arabic_font
    )
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(arabic_text(footer_text), footer_style))
    
    doc.build(elements)
    
    buffer.seek(0)
    
    # تسجيل النشاط
    log_activity('طباعة تقرير', 'جدول', None, f'تم طباعة تقرير جدول المعلمين')
    
    return send_file(buffer, as_attachment=True, download_name=f'schedule_table_{datetime.now().strftime("%Y%m%d")}.pdf', mimetype='application/pdf')

# تعديل شهادة للمدير
@admin_bp.route('/certificates/edit/<int:cert_id>', methods=['POST'])
@login_required
def edit_certificate(cert_id):
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    certificate = Certificate.query.get_or_404(cert_id)
    
    try:
        certificate.student_name = request.form.get('student_name', certificate.student_name)
        certificate.nationality = request.form.get('nationality', certificate.nationality)
        certificate.phone = request.form.get('phone', certificate.phone)
        certificate.expected_completion_date = datetime.strptime(
            request.form.get('expected_completion_date'), '%Y-%m-%d'
        ).date()
        certificate.narration_type = request.form.get('narration_type', certificate.narration_type)
        certificate.halaqah = request.form.get('halaqah', certificate.halaqah)
        certificate.completion_type = request.form.get('completion_type', certificate.completion_type)
        certificate.teacher_name = request.form.get('teacher_name', certificate.teacher_name)
        certificate.notes = request.form.get('notes', certificate.notes)
        certificate.updated_by = current_user.id
        certificate.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        # تسجيل النشاط
        log_activity('تعديل شهادة', 'certificate', cert_id, f'تم تعديل شهادة الطالب {certificate.student_name}')
        
        flash('تم تحديث الشهادة بنجاح', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('حدث خطأ أثناء تحديث الشهادة', 'danger')
    
    return redirect(url_for('certificates.admin_manage'))

# إدارة رصيد الإجازات
@admin_bp.route('/leave_balance')
@login_required
def leave_balance_management():
    if not admin_required():
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    employees = User.query.filter_by(role=Role.EMPLOYEE).order_by(User.name).all()
    return render_template('admin/leave_balance.html', employees=employees)

# تحديث رصيد الإجازات
@admin_bp.route('/leave_balance/update/<int:user_id>', methods=['POST'])
@login_required
def update_leave_balance(user_id):
    if not admin_required():
        return jsonify({'success': False, 'message': 'غير مصرح'}), 403
    
    user = User.query.get_or_404(user_id)
    
    try:
        new_balance = int(request.json.get('balance', 0))
        if new_balance < 0:
            return jsonify({'success': False, 'message': 'الرصيد لا يمكن أن يكون سالباً'}), 400
        
        old_balance = user.leave_balance
        user.leave_balance = new_balance
        db.session.commit()
        
        # تسجيل النشاط
        log_activity('تحديث رصيد إجازة', 'user', user_id, 
                    f'تم تحديث رصيد إجازة {user.name} من {old_balance} إلى {new_balance}')
        
        return jsonify({
            'success': True,
            'message': 'تم تحديث الرصيد بنجاح',
            'new_balance': new_balance
        })
    except ValueError:
        return jsonify({'success': False, 'message': 'قيمة غير صحيحة'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'حدث خطأ أثناء تحديث الرصيد'}), 500
